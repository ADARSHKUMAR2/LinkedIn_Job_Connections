import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.core.database import SessionLocal
from app.models.job import JobListing
from app.models.connection import LinkedInConnection
from app.services.ai_matcher import calculate_matches

def export_matches_to_excel():
    print("==================================================================")
    print("📊 Compiling referral matches into styled target_jobs.xlsx...")
    print("==================================================================")

    db = SessionLocal()
    export_path = os.path.join(os.path.dirname(__file__), "../../target_jobs.xlsx")

    try:
        
        matched_rows = calculate_matches(db)

        if not matched_rows:
            print("⚠️  No relational match vectors found in the database to export yet.")
            return

        # 3. Use Pandas to normalize arrays into a clean DataFrame frame
        df = pd.DataFrame(matched_rows)
        
        # Write to excel using standard pandas engine
        df.to_excel(export_path, index=False, sheet_name="Target Matches")

        # 4. Use OpenPyXL to inject professional styling and sheet layouts
        wb = load_workbook(export_path)
        ws = wb["Target Matches"]

        # Design system tokens
        navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        white_bold_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Arial", size=10)
        link_font = Font(name="Arial", size=10, color="0000FF", underline="single")
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        # Format Header Row
        for cell in ws[1]:
            cell.fill = navy_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 26

        # Format Data Rows
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                # Turn the Application Link column into actual clickable sheet hyperlinks
                if ws.cell(row=1, column=col).value == "Application Link" and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = link_font

        # Auto-fit columns to clear layout text bounding blocks cleanly
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(export_path)
        print(f"✅ SUCCESS: Formatted spreadsheet generated cleanly at:\n   👉 {os.path.abspath(export_path)}")

    except Exception as e:
        print(f"❌ EXCEL BUILDER FAILED: {e}")
    finally:
        db.close()
        print("==================================================================")

if __name__ == "__main__":
    export_matches_to_excel()